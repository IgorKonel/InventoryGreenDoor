import os  # Для открытия Excel файла в конце работы

import variables

from windows.window import get_window  # Импорт файла с функцией создания окна

import json  # Для базы данных

from creds import api_key  # Импорт из пакета API key

from googleapiclient.discovery import build  # Библиотеки для работы Google Sheets API

import openpyxl  # Библиотека для работы с Excel
from openpyxl.styles import Font, Alignment  # Шрифты и расположение текста

from datetime import datetime as dt  # Работа с месяцами для проверки нужно ли обновить ссылку

NAME_LIST = "Invent WD/ZD"


def print_list(list_resp, str_resp):
    """
        Ввывод
    """
    print(f'\n{str_resp} \n')
    for pos, num in list_resp:
        print(f'{pos:<60} {num:>8}')


def print_dict_list(list_resp, str_resp):
    print(f'\n{str_resp} \n')
    for pos in list_resp:
        if len(pos) == 4:
            print(f'{pos[0]:<60} {pos[1]:<9} {pos[2]:<11} {pos[3]:>3}')
        elif len(pos) == 2:
            print(f'{pos[0]:<60} {pos[1]:<9}')


def get_service_simple():
    """
    Чтение таблиц без возможности их редактирования
    :return:
    """
    return build('sheets', 'v4', developerKey=api_key)


def get_resp(range_list):
    path = NAME_LIST + '!' + range_list
    return sheet.values().batchGet(spreadsheetId=sheet_id, ranges=[path]).execute()['valueRanges'][0]['values']


def get_list_order(range_list):
    return [s for s in range_list if s[len(range_list[0]) - 1] != '0']


service = get_service_simple()  # С использованием API
sheet = service.spreadsheets()

"""
************************************************************************************************************************
                                        Проверка на актуальность таблицы
************************************************************************************************************************
"""

# Открываем файл, в котором хранятся данные о номере месяца последнего изменения и id таблицы
with open('data.json') as json_file:
    data = json.load(json_file)

    # Данные из файла
    print('Текущие данные:')
    print(f'month : {data["month"]}')  # Номер месяца
    print(f'sheet_id : {data["sheet_id"]}')  # Id таблицы

    month_current = dt.now().month  # Текущий месяц

    # Сравниваем на неравенство текущий месяц и с последнего изменения
    if data["month"] != month_current:

        get_window()  # Запускаем окно из файла windows.window.py
        # Здесь мы получаем новое значение для variables.SHEET_ID

        # TODO: Сделать обработку строки, чтобы из ссылки получалось id таблицы
        sheet_id = variables.SHEET_ID
        sheet_id = sheet_id[39:]
        sheet_id = sheet_id[:sheet_id.index('/')]

        data['month'] = month_current  # Заносим новые данные в словарь
        data['sheet_id'] = sheet_id

        with open('data.json', 'w') as outfile:
            json.dump(data, outfile)  # Запись словаря в JSON-file
    else:
        sheet_id = data["sheet_id"]  # Подгружаем sheet_id из JSON

    print('Данные для работы:')
    print(f'sheet_id now : {sheet_id}')  # Контрольный вывод id таблицы
    print(f'month : {month_current}')  # Контрольный вывод номера месяца, когда было последнее изменение

"""
************************************************************************************************************************
                                             Запрос на получение данных
************************************************************************************************************************
"""

# Получаем списки:
resp_house = get_resp('AP5:AQ48')  # Хоз-товары на всех
resp_tea = get_resp('AP49:AQ85')  # Чаи на всех
resp_art_count = get_resp('B5:C48')  # Артикулы и кол-во

# Wooden Door
resp_wd_pos = get_resp('AP5:AP85')
order_wd = get_resp('AR5:AR85')
order_wdl = get_resp('AS5:AS85')

"""
************************************************************************************************************************
                                             Редактирование списков
************************************************************************************************************************
"""

# Хоз-товары
for i in range(len(resp_house)):
    resp_house[i].insert(1, resp_art_count[i][0])  # Вставляем артикулы
    resp_house[i].insert(2, resp_art_count[i][1])  # Вставляем Шт\Уп

# Добавляем
for i in range(len(resp_wd_pos)):
    order_wd[i].insert(0, resp_wd_pos[i][0])
    order_wdl[i].insert(0, resp_wd_pos[i][0])

print_list(order_wd, 'Лубянка')



# print_list(resp_house, 'Список общих хоз-товаров:')
# print_list(resp_tea, 'Список общих чаёв:')
# print_list(resp_art_count, 'Список артикулов и кол-ва')

# TODO:
# Формируем конечные списки на закупку

# Убираем позиции с нулевым значением с помощью функции get_list_order

# Общие
order_house = get_list_order(resp_house)
order_tea = get_list_order(resp_tea)

# Wooden Door
order_wd = get_list_order(order_wd)
order_wdl = get_list_order(order_wdl)

"""
************************************************************************************************************************
                                                    Работа с Excel
************************************************************************************************************************
"""

# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title='Инвентаризация', index=0)

# получаем лист, с которым будем работать
sheet = wb['Инвентаризация']

#  Подгоняем по ширине столбцы
sheet.column_dimensions['A'].width = 65
sheet.column_dimensions['B'].width = 11
sheet.column_dimensions['C'].width = 14
sheet.column_dimensions['D'].width = 5

#  Запись хозтоваров
sheet.cell(row=1, column=1).value = 'Список хозтоваров:'
sheet.cell(row=1, column=1).font = Font(size=25)

count = 3

for row in range(count, len(order_house)+count):
    for col in range(1, 5):
        if col == 4:
            sheet.cell(row=row, column=col).value = int(order_house[row - 3][col - 1])
        else:
            sheet.cell(row=row, column=col).value = order_house[row-3][col-1]
        if col > 1:
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')

count += len(order_house) + 2


#  Запись чаёв
sheet.cell(row=count, column=1).value = 'Список чаёв:'
sheet.cell(row=count, column=1).font = Font(size=25)
count += 2

for row in range(count, len(order_tea)+count):
    for col in range(1, 3):
        if col == 2:
            sheet.cell(row=row, column=col).value = int(order_tea[row - count][col - 1])
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        else:
            sheet.cell(row=row, column=col).value = order_tea[row-count][col-1]
count += len(order_tea) + 2

# Список что отдать Лубянке
sheet.cell(row=count, column=1).value = 'Отдать Лубянке:'
sheet.cell(row=count, column=1).font = Font(size=25)
count += 2

for row in range(count, len(order_wd)+count):
    for col in range(1, 3):
        if col == 2:
            sheet.cell(row=row, column=col).value = int(order_wd[row - count][col - 1])
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        else:
            sheet.cell(row=row, column=col).value = order_wd[row-count][col-1]
count += len(order_wd) + 2

# Список что отдать Сухарю
sheet.cell(row=count, column=1).value = 'Отдать Сухарю:'
sheet.cell(row=count, column=1).font = Font(size=25)
count += 2

for row in range(count, len(order_wdl)+count):
    for col in range(1, 3):
        if col == 2:
            sheet.cell(row=row, column=col).value = int(order_wdl[row - count][col - 1])
            sheet.cell(row=row, column=col).alignment = Alignment(horizontal='center')
        else:
            sheet.cell(row=row, column=col).value = order_wdl[row-count][col-1]
count += len(order_wdl) + 2


#  Сохранение в файл
wb.save('Order.xlsx')

#  Открытие файла (файл должен быть закрыт)
os.startfile('Order.xlsx')